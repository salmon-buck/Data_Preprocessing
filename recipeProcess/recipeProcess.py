import os
from openpyxl import Workbook
from Data_Preprocessing.recipeProcess.preProcessing import preProcessing
import copy

alpha=50
beta=0.2
def tf_at_document(string) :
    term_freq={}
    word_list=string.split()
    word_set=set(word_list)

    for w in word_set :
        term_freq[w] = word_list.count(w)                    # tf(t, d) 계산
    return term_freq

def prob_at_collection(total_wordset, total_wordlist) :     # P(t|Mc)
    collection_prob={}
    length=len(total_wordlist)
    for w in total_wordset :
        collection_prob[w] = total_wordlist.count(w)/length
    return collection_prob

def smoothing(t_f, t_f2, collection_prob) :
    smoothing_prob={}
    term_freq=copy.deepcopy(t_f)
    term_freq2=copy.deepcopy(t_f2)
    for food in term_freq.keys() :
        dic = {}
        length=sum(term_freq[food].values())
        length2=sum(term_freq2[food].values())

        for word in term_freq[food].keys() :
            dic[word]=(term_freq[food][word] + alpha*collection_prob[word]) / (length+alpha)
            if word in term_freq2[food].keys() :
                dic[word]+=(term_freq2[food][word] + alpha * collection_prob[word]) / (length2 + alpha) * beta
                del term_freq2[food][word]

        for w in term_freq2[food].keys() :
            dic[w] = (term_freq2[food][w] + alpha * collection_prob[w]) / (length2 + alpha) * beta

        smoothing_prob[food]=dic
    return smoothing_prob


f = open('recipe.txt', 'rt', encoding='UTF8')
list = []

input_path = os.getcwd()
txt_path=os.path.join(input_path, '../../Data_Preprocessing/processedRecipe/')

for i in range(100):
    line = f.readline()
    lines = line.split(';')
    # print(i)

    description = preProcessing(lines[2])
    ingredient = preProcessing(lines[4])

    lines[2] = ",".join(description)
    lines[4] = ",".join(ingredient)

    filename = str(lines[1]) + ".txt"

    list.append(filename)

    # print(filename)
    if os.path.isfile(filename):
        df = open(filename, 'rt', encoding='UTF8')
        des = ""
        for j in range(2):
            des = df.readline()
        lines[2] = lines[2] + "," + des[:-1]
        df.close()

    nf = open(txt_path + filename, 'wt', encoding='UTF8')

    for j in range(1, len(lines)):
        nf.writelines(lines[j])
        nf.write('\n')
    nf.close()

f.close()

f = open('recipe.txt', 'rt', encoding='UTF8')
list = []
db = {}
term_freq = {}
term_freq2={}
total_wordset=set()
total_wordlist=[]
for i in range(100):
    dic = {}

    line = f.readline()
    lines = line.split(';')
    # print(i)

    rname = str.strip(lines[1])
    #rname = ' '.join(preProcessing(str.strip(rname)))

    #description = preProcessing(lines[2])
    #ingredient = preProcessing(lines[4])

    dic['name'] = rname
    dic['description'] = str.strip(lines[2])
    dic['country'] = str.strip(lines[3])
    dic['ingredient'] = str.strip(lines[4])
    dic['recipe'] = str.strip(lines[5])
    dic['time'] = str.strip(lines[6])
    dic['ImageUrl'] = str.strip(lines[7])
    db[rname] = dic

    dec_ing_rcp = ' '.join(preProcessing(dic['name'])) + ' ' +  ' '.join(preProcessing(dic['description'])) + ' ' + ' '.join(preProcessing(dic['ingredient']))
    term_freq[rname] = tf_at_document(dec_ing_rcp)
    term_freq2[rname] = tf_at_document(' '.join(preProcessing(dic['recipe'])))

    total_wordset = total_wordset.union((dec_ing_rcp+' ' + ' '.join(preProcessing(dic['recipe']))).split())
    total_wordlist.extend((dec_ing_rcp+' ' + ' '.join(preProcessing(dic['recipe']))).split())

f.close()

collection_prob={}
collection_prob=prob_at_collection(total_wordset, total_wordlist)                     # P(t|Mc)


smoothing_prob={}
smoothing_prob=smoothing(term_freq, term_freq2, collection_prob)


############################################################ 엑셀 DB ############################################################

write_wb = Workbook()                                                           # workbook 생성
sheets=[]                                                                       # sheet 이름 저장해놓는 sheets
for i in range(4):
    if i == 0:                                                                  # 디폴트로 생성되는 sheet1이 있기때문에 이름을 filenames[0] 의 이름으로 바꾼다
        s = write_wb.active
        s.title = 'DB'

    elif i == 1:
        s = write_wb.create_sheet('P(t|d)')

    elif i == 2:
        s = write_wb.create_sheet('P(t|C)')

    else :
        s = write_wb.create_sheet('Smoothing')
    sheets.append(s)
    # 디폴트로 생성되는 sheet1이 있기때문에 이름을 filenames[0] 의 이름으로 바꾼다


db=sorted(db.items(), key=lambda x : x[1]['country'])
count=1
for word in db :                                                                  # DB 시트
    sheets[0].cell(count, 1, word[1]['name'])                                     # 1열 : 요리이름
    sheets[0].cell(count, 2, word[1]['description'])                              # 2열 : 설명
    sheets[0].cell(count, 3, word[1]['country'])                                  # 3열 : 국가
    sheets[0].cell(count, 4, word[1]['ingredient'])                               # 4열 : 재료
    sheets[0].cell(count, 5, word[1]['recipe'])                                   # 5열 : 레시피 순서
    sheets[0].cell(count, 6, word[1]['time'])                                     # 6열 : 요리시간
    sheets[0].cell(count, 7, word[1]['ImageUrl'])                                 # 7열 : 이미지주소

    count+=1

count=1
cnt=1
for food in term_freq.keys() :                                                    # P(t|d) & Smoothing 시트
    sheets[1].cell(count, 1, food)
    sheets[3].cell(count, 1, food)

    for word in term_freq[food].keys() :
        sheets[1].cell(count, 2, word)
        sheets[1].cell(count, 3, term_freq[food][word])
        sheets[3].cell(count, 2, word)
        sheets[3].cell(count, 3, smoothing_prob[food][word])
        count+=1

    count+=1
    for w in term_freq2[food].keys() :
        sheets[1].cell(cnt, 4, w)
        sheets[1].cell(cnt, 5, term_freq2[food][w])
        cnt += 1

    cnt+=1
    if count>cnt :
        cnt=count
    else :
        count=cnt

count=1
for word in collection_prob.keys() :                                              # P(t|C) 시트
    sheets[2].cell(count, 1, word)
    sheets[2].cell(count, 2, collection_prob[word])
    count+=1

write_wb.save('DB.xlsx')     # 엑셀파일에 저장
